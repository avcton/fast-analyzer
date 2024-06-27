import re
import openpyxl
import pandas as pd
from datetime import datetime, timedelta


class TimeTable():

    # Function to extract proper start time from date time object
    def _extract_start_time(time: str, delta: int):
        _time = datetime.strptime(time, '%I:%M %p')
        start = _time - timedelta(minutes=delta)
        return start.strftime('%I:%M %p')
    
    # Function to clean the course title
    def _clean_course_title(input_string: str):
        pattern = re.compile(r'\((.*?)\)')

        # Find all text within parentheses
        matches = pattern.findall(input_string)

        # Iterate through all matches to clean them
        for match in matches:
            # Remove hyphens
            cleaned_match = match.replace('-', ' ')
            # Add space after commas if not already present
            cleaned_match = re.sub(r',([^ ])', r', \1', cleaned_match)
            # Replace the original match with the cleaned one in the input string
            input_string = input_string.replace(f'({match})', f'({cleaned_match})')

        return input_string.strip()

    # Function to extract section part from course name
    def _extract_section(course: str):
        # Find text within parentheses
        start_index = course.find('(')
        end_index = course.find(')')
        if start_index != -1 and end_index != -1:
            return course[start_index:end_index + 1]
        else:
            return None

    def __init__(self, title: str = None) -> None:
        self.title = title
        self.data = None
        self.loaded = False
        self.options = []
        # Define regex pattern to match time formats with optional periods
        self.timePattern = r'(\d{1,2}):(\d{2})(?:\.|)(am|pm)\.?'
        # Determine time interval from one cell to another in the sheet
        # Can be hardcoded by observation
        # Can also be determined by the following formula:
        # Lecture time for any course in minutes / No. of merged cells occupied by the lecture in the sheet
        self.delta = 80 // 8  # Lecture of 80 mins / 8 cells it occupies

    def load(self, file: any):
        try:
            self.file = file

            if self.title == None:
                self.title = pd.read_excel(
                    file, header=None, nrows=1).bfill(axis=1).iloc[0, 0]

            self.batches = pd.read_excel(file, header=None, skiprows=1, nrows=1).iloc[0, 1:].dropna(
                ignore_index=True).apply(lambda batch: batch.strip())

            # Reading time intervals
            time_series = pd.read_excel(file, header=None, skiprows=3, nrows=1)
            time_series_range = len(time_series.columns) - 1
            time_series = time_series.bfill(axis=1).iloc[0, [2]].apply(str)
            time_series.name = "Time Series"

            # Use re.search to find the pattern in the input string
            # Using IGNORECASE for case insensitivity
            match = re.search(self.timePattern,
                              time_series.iat[0], re.IGNORECASE)

            if match:
                hour = int(match.group(1))
                minute = match.group(2)
                period = match.group(3).upper()  # Convert am/pm to uppercase

                # Adjust hour if necessary (e.g., convert from 12-hour to 24-hour format)
                if period == 'PM' and hour < 12:
                    hour += 12
                elif period == 'AM' and hour == 12:
                    hour = 0

                # Format the hour and minute as '08:00 AM'
                time_series.iat[0] = f'{hour:02}:{minute} {period}'

            else:
                # Fall Back to Hardcoded Expected Time
                time_series.iat[0] = '08:00 AM'

            # Parse starting time
            time_series.iat[0] = start_time = datetime.strptime(
                time_series.iat[0], '%I:%M %p') + timedelta(minutes=self.delta)

            # Generate time values with 10-minute intervals
            current_time = start_time
            current_index = time_series.index[0] + 1
            while current_index < time_series_range:
                current_time += timedelta(minutes=self.delta)
                time_series[current_index] = current_time
                current_index += 1

            time_series = time_series.apply(
                lambda time: time.strftime('%I:%M %p'))

            # Manifesting Timetable Time Interval Row
            timetable_columns = time_series.copy()
            timetable_columns[0] = 'Day'
            timetable_columns[1] = 'Room'
            timetable_columns = timetable_columns.sort_index()

            # Loading the Timetable
            timetable = pd.read_excel(file, engine='openpyxl', dtype=str)

            # Load the workbook with openpyxl to access merged cells and deal with them
            # This potential workbook loading for openpyxl is a little time taking though: estimated time = 6 seconds
            workbook = openpyxl.load_workbook(file, data_only=True)
            sheet = workbook.active

            # Extract merged cell ranges
            merged_ranges = sheet.merged_cells.ranges

            # Update the DataFrame with merged cell values
            for merged_range in merged_ranges:
                # Get the boundaries of the merged range
                min_row = merged_range.min_row
                min_col = merged_range.min_col
                max_row = merged_range.max_row
                max_col = merged_range.max_col

                value = sheet.cell(min_row, min_col).value

                # Fill the merged range in the DataFrame
                for row in range(min_row, max_row + 1):
                    for col in range(min_col, max_col + 1):
                        if row-2 >= 0 and row-2 < timetable.shape[0] and col-1 >= 0 and col - 1 < timetable.shape[1]:
                            timetable.iat[row-2, col-1] = value

            # Melting Timetable and building a refined version
            timetable = timetable.iloc[3:, :-1]
            timetable.columns = timetable_columns
            timetable['Day'] = timetable['Day'].ffill().apply(
                lambda x: x.split()[0])
            days_order = timetable['Day'].unique()

            timetable_refined = timetable.melt(id_vars=['Day', 'Room'],
                                               var_name='Time', value_name='Course')
            timetable_refined.dropna(subset=['Course'], inplace=True)
            timetable_refined['Day'] = pd.Categorical(timetable_refined['Day'],
                                                      categories=days_order, ordered=True)
            timetable_refined = timetable_refined.pivot_table(
                index=['Day', 'Room', 'Course'], values='Time', aggfunc=lambda x: ' - '.join(x), observed=True).reset_index()
            
            # Check if the data is correct
            assert timetable_refined['Time'].apply(len).mean() > 20, "Incorrect Data Format"

            timetable_refined['Time'] = timetable_refined['Time'].apply(
                lambda x: TimeTable._extract_start_time(x.split(' - ')[0], self.delta) + ' - ' + x.split(' - ')[-1])

            # Sorting by Start Time
            timetable_refined['Start Time'] = timetable_refined['Time'].apply(
                lambda x: x.split(' - ')[0])
            timetable_refined['Start Time'] = pd.to_datetime(
                timetable_refined['Start Time'], format='%I:%M %p')
            timetable_refined.sort_values(
                by=['Day', 'Start Time', 'Course', 'Room'], inplace=True)
            timetable_refined.drop(columns='Start Time', inplace=True)

            timetable_refined['Course'] = timetable_refined['Course'].apply(
                TimeTable._clean_course_title)

            self.columns = ['Day', 'Time', 'Room', 'Course']
            timetable_refined = timetable_refined[self.columns]

            self.options = set(timetable_refined['Course'].to_list())
            self.data = timetable_refined
            self.loaded = True

        except Exception as e:
            print(f"Caught an error: {e}")
            self.loaded = False

        finally:
            return self.loaded

    def readByCourse(self, courses: list[str]):
        assert (self.loaded == True)
        result = self.data[self.data.isin(
            {'Course': courses}).any(axis=1)].copy().reset_index(drop=True)

        result['Start Time'] = result['Time'].apply(
            lambda x: x.split(' - ')[0])
        result['Start Time'] = pd.to_datetime(
            result['Start Time'], format='%I:%M %p')

        # Apply function to extract section for each course
        result['Section'] = result['Course'].apply(TimeTable._extract_section)

        if result['Section'].nunique() == 1:
            section_to_remove = result.loc[0, 'Section']
            result['Course'] = result['Course'].apply(
                lambda x: x.replace(section_to_remove, '').strip())

        # Drop the 'Section' column as it's no longer needed
        result.drop(columns=['Section'], inplace=True)

        result.sort_values(
            by=['Day', 'Start Time', 'Course', 'Room'], inplace=True)

        result.drop(columns='Start Time', inplace=True)

        # Find unique days
        unique_days = result['Day'].unique()

        result_refined = pd.DataFrame(columns=result.columns, data={
                                    col: [''] for col in result.columns})

        # Iterate through the DataFrame and insert an empty row after each unique day
        for day in unique_days:
            day_df = result[result['Day'] == day]
            result_refined = pd.concat([result_refined, day_df], ignore_index=True)
            # Append an empty row after the current day's rows
            empty_row = pd.DataFrame({col: [''] for col in result.columns})
            result_refined = pd.concat([result_refined, empty_row], ignore_index=True)

        result_refined = result_refined.iloc[1:-1, :]

        return result_refined
